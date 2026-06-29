from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import os
import time

app = FastAPI(
    title="Gastos Mensuales API",
    description="API para leer el Excel de gastos y exponerlo a Power BI",
    version="1.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "Gastos_Mensuales_v3.xlsx")

_cache = {"data": None, "ts": 0}

def load_excel() -> dict[str, pd.DataFrame]:
    now = time.time()
    if _cache["data"] and (now - _cache["ts"]) < 600:
        return _cache["data"]
    if not os.path.exists(EXCEL_PATH):
        raise HTTPException(status_code=500, detail=f"Archivo Excel no encontrado en: {EXCEL_PATH}")
    sheets = pd.read_excel(EXCEL_PATH, sheet_name=None, header=1)
    _cache["data"] = sheets
    _cache["ts"] = now
    return sheets

def sheet_to_records(df: pd.DataFrame) -> list[dict]:
    df = df.dropna(how="all").reset_index(drop=True)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.where(pd.notna(df), other=None)
    return df.to_dict(orient="records")

@app.get("/")
def root():
    return {"api": "Gastos Mensuales API", "endpoints": ["/panel","/tarjetas","/personales","/hogar","/deudas","/referencia","/resumen","/objetivos"]}

@app.get("/panel")
def get_panel():
    sheet = load_excel().get("Panel Principal")
    if sheet is None: raise HTTPException(status_code=404, detail="Pestaña no encontrada")
    return {"sheet": "Panel Principal", "data": sheet_to_records(sheet)}

@app.get("/tarjetas")
def get_tarjetas():
    sheet = load_excel().get("Tarjetas_Bancos")
    if sheet is None: raise HTTPException(status_code=404, detail="Pestaña no encontrada")
    return {"sheet": "Tarjetas_Bancos", "data": sheet_to_records(sheet)}

@app.get("/personales")
def get_personales():
    sheet = load_excel().get("Gastos_Personales")
    if sheet is None: raise HTTPException(status_code=404, detail="Pestaña no encontrada")
    return {"sheet": "Gastos_Personales", "data": sheet_to_records(sheet)}

@app.get("/hogar")
def get_hogar():
    sheet = load_excel().get("Hogar")
    if sheet is None: raise HTTPException(status_code=404, detail="Pestaña no encontrada")
    return {"sheet": "Hogar", "data": sheet_to_records(sheet)}

@app.get("/deudas")
def get_deudas():
    sheet = load_excel().get("Deudas_Varios")
    if sheet is None: raise HTTPException(status_code=404, detail="Pestaña no encontrada")
    return {"sheet": "Deudas_Varios", "data": sheet_to_records(sheet)}

@app.get("/referencia")
def get_referencia():
    sheet = load_excel().get("Referencia")
    if sheet is None: raise HTTPException(status_code=404, detail="Pestaña no encontrada")
    return {"sheet": "Referencia", "data": sheet_to_records(sheet)}

@app.get("/objetivos")
def get_objetivos():
    sheets = load_excel()
    sheet = sheets.get("Objetivos")
    if sheet is None: raise HTTPException(status_code=404, detail="Pestaña 'Objetivos' no encontrada")

    # Leer con header correcto (fila 5 del Excel = fila 3 con header=1)
    import io
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Objetivos"]

    # Extraer supuesto ahorro mensual (B3)
    ahorro_mensual = ws["B3"].value or 200000

    # Extraer objetivos (filas 6-8, cols A-I)
    objetivos = []
    for row in ws.iter_rows(min_row=6, max_row=8, min_col=1, max_col=9, values_only=True):
        if row[0] and str(row[0]).strip():
            meta = row[3] or 0
            ahorrado = row[4] or 0
            faltante = max(meta - ahorrado, 0)
            pct = round(ahorrado / meta * 100, 1) if meta > 0 else 0
            meses = -(-faltante // ahorro_mensual) if ahorro_mensual > 0 and faltante > 0 else 0
            objetivos.append({
                "objetivo": row[0],
                "prioridad": row[1],
                "fecha_limite": str(row[2]) if row[2] else "",
                "monto_meta": meta,
                "ya_ahorrado": ahorrado,
                "faltante": faltante,
                "pct_avance": pct,
                "meses_restantes": meses,
                "ahorro_mensual": ahorro_mensual
            })

    return {"ahorro_mensual": ahorro_mensual, "objetivos": objetivos}

@app.get("/resumen")
def get_resumen():
    panel = load_excel().get("Panel Principal")
    if panel is None: raise HTTPException(status_code=404, detail="Pestaña no encontrada")
    ROWS_OF_INTEREST = [
        "Tarjetas & Bancos", "Gastos Personales", "Hogar", "Deudas & Varios",
        "TOTAL GASTOS", "Ingreso Franco", "Ingreso Abi", "Pago de negocio",
        "Ingreso por rendimientos", "Nos deben", "TOTAL INGRESOS",
        "¿LLEGAMOS A PAGAR?", "% Franco", "% Abi",
    ]
    panel = panel.dropna(how="all").reset_index(drop=True)
    label_col = panel.columns[0]
    month_cols = panel.columns[1:]
    records = []
    for _, row in panel.iterrows():
        label = str(row[label_col]).strip()
        if label not in ROWS_OF_INTEREST:
            continue
        for mes in month_cols:
            v = row[mes]
            records.append({
                "categoria": label,
                "mes": str(mes).strip(),
                "valor": None if pd.isna(v) else v
            })
    return {"data": records}

@app.get("/health")
def health():
    return {"status": "ok"}
